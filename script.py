import json
import os
import re
import time

import openpyxl
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# ─── Configuration ────────────────────────────────────────────────────────────

load_dotenv()

ACTION_DELAY = float(os.getenv("ACTION_DELAY", "1.0"))
PAGE_TIMEOUT = int(os.getenv("PAGE_TIMEOUT", "30"))

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "TVA A TRANSFERER.xlsx")
PROGRESS_DIR = os.path.dirname(__file__)

SHEETS_TO_PROCESS = ["TVA 3", "TVA 4", "TVA 5 ", "TVA 6 ", "TVA 7", "TVA 8"]

BASE_URL = "https://cfspro.impots.gouv.fr"
DELEGATION_URL = f"{BASE_URL}/opale_usager/SaisieSirenDelegation.do?choixSirenIn=true"

# Services à déléguer dans l'ordre (le texte exact des labels sur le site)
SERVICES = [
    {"label": "Messagerie",                "check_all": False},
    {"label": "Déclarer TVA",              "check_all": False},
    {"label": "Payer TVA",                 "check_all": False},
    {"label": "Consulter le Compte fiscal", "check_all": True},
    {"label": "Déclarer le Résultat",      "check_all": False},
]


# ─── Utilitaires ──────────────────────────────────────────────────────────────

def progress_file_for(sheet_name):
    """Retourne le chemin du fichier de progression pour un onglet."""
    safe_name = sheet_name.strip().replace(" ", "_")
    return os.path.join(PROGRESS_DIR, f"progress_{safe_name}.json")


def load_progress(sheet_name):
    """Charge la progression depuis le fichier JSON pour un onglet."""
    pf = progress_file_for(sheet_name)
    if os.path.exists(pf):
        with open(pf, "r") as f:
            return json.load(f)
    return {"siren_index": 0}


def save_progress(sheet_name, siren_index):
    """Sauvegarde la progression actuelle pour un onglet."""
    pf = progress_file_for(sheet_name)
    with open(pf, "w") as f:
        json.dump({"siren_index": siren_index}, f, indent=2)


def clear_progress(sheet_name):
    """Supprime le fichier de progression pour un onglet."""
    pf = progress_file_for(sheet_name)
    if os.path.exists(pf):
        os.remove(pf)


def extract_abonne_number(cell_value):
    """Extrait le numéro d'abonné depuis 'ABONNE 20260410001818'."""
    match = re.search(r"\d+", str(cell_value))
    return match.group() if match else None


def read_excel():
    """Lit le fichier Excel et retourne les données par onglet."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    data = {}
    for sheet_name in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            print(f"  Onglet '{sheet_name}' non trouve, ignore.")
            continue
        ws = wb[sheet_name]
        abonne = extract_abonne_number(ws["A1"].value)
        sirens = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=4, max_col=4):
            cell = row[0]
            if cell.value is not None:
                try:
                    siren = str(int(float(str(cell.value)))).zfill(9)
                    sirens.append(siren)
                except (ValueError, TypeError):
                    print(f"    Valeur invalide ignoree en {cell.coordinate}: {cell.value}")
                    continue
        data[sheet_name] = {"abonne": abonne, "sirens": sirens}
        print(f"  {sheet_name}: abonne={abonne}, {len(sirens)} SIRENs")
    return data


# ─── Selenium : actions sur le navigateur ─────────────────────────────────────

def wait_and_find(driver, selector, by=By.CSS_SELECTOR, timeout=None):
    """Attend qu'un élément soit présent et le retourne."""
    timeout = timeout or PAGE_TIMEOUT
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )


def wait_and_click(driver, selector, by=By.CSS_SELECTOR, timeout=None):
    """Attend qu'un élément soit cliquable et clique dessus."""
    timeout = timeout or PAGE_TIMEOUT
    element = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )
    element.click()
    time.sleep(ACTION_DELAY)
    return element


def fill_input(driver, selector, value, by=By.CSS_SELECTOR):
    """Remplit un champ input."""
    element = wait_and_find(driver, selector, by)
    element.clear()
    element.send_keys(value)
    time.sleep(ACTION_DELAY / 2)
    return element


def init_driver():
    """Initialise le navigateur Chrome."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def navigate_to_delegation_page(driver, max_retries=3):
    """
    Depuis la page d'accueil ou la page services, navigue jusqu'à la page SIREN.
    Gère la page d'erreur "Fermer", les nouveaux onglets et la fenêtre popup.
    """
    for attempt in range(max_retries):
        try:
            print(f"    Tentative {attempt + 1}/{max_retries}...")
            print(f"    URL actuelle: {driver.current_url}")

            # Vérifier si on est sur une page d'erreur avec bouton "Fermer"
            try:
                fermer_link = driver.find_element(By.XPATH, "//a[contains(text(),'Fermer')]")
                print("    Page d'erreur detectee, clic sur 'Fermer'...")
                fermer_link.click()
                time.sleep(ACTION_DELAY * 2)
            except Exception:
                pass

            # Fermer toutes les fenêtres/onglets sauf le premier et y revenir
            main_handle = driver.window_handles[0]
            for handle in driver.window_handles[1:]:
                driver.switch_to.window(handle)
                driver.close()
            driver.switch_to.window(main_handle)
            time.sleep(ACTION_DELAY)

            # S'assurer qu'on est sur la page d'accueil
            if "accueil" not in driver.current_url:
                driver.get(f"{BASE_URL}/mire/accueil.do")
                time.sleep(ACTION_DELAY * 2)

            # Étape 1 : Cliquer sur "Gérer les services" (ouvre un nouvel onglet)
            print("    Clic sur 'Gerer les services'...")
            handles_before = set(driver.window_handles)
            wait_and_click(driver, "//a[contains(text(),'rer les services')]", by=By.XPATH, timeout=PAGE_TIMEOUT)
            time.sleep(ACTION_DELAY * 3)

            # Attendre que le nouvel onglet/fenêtre apparaisse et y basculer
            WebDriverWait(driver, PAGE_TIMEOUT).until(lambda d: len(d.window_handles) > len(handles_before))
            new_handles = set(driver.window_handles) - handles_before
            services_handle = new_handles.pop()
            driver.switch_to.window(services_handle)
            print(f"    Page services, URL: {driver.current_url}")
            time.sleep(ACTION_DELAY * 2)

            # Étape 2 : Ouvrir la page de délégation SIREN
            # Le lien utilise javascript:winPop() qui fait window.open()
            # On l'exécute directement via JS pour ouvrir le popup
            print("    Ouverture du popup delegation via JS...")
            handles_before_popup = set(driver.window_handles)
            driver.execute_script(
                "window.open('SaisieSirenDelegation.do?choixSirenIn=true','delegation','width=810,height=600');"
            )
            time.sleep(ACTION_DELAY * 3)

            # Attendre que la fenêtre popup apparaisse et y basculer
            WebDriverWait(driver, PAGE_TIMEOUT).until(lambda d: len(d.window_handles) > len(handles_before_popup))
            popup_handles = set(driver.window_handles) - handles_before_popup
            popup_handle = popup_handles.pop()
            driver.switch_to.window(popup_handle)
            print(f"    Popup delegation, URL: {driver.current_url}")
            time.sleep(ACTION_DELAY)

            # Vérifier si on a atterri sur la page d'erreur
            try:
                driver.find_element(By.XPATH, "//a[contains(text(),'Fermer')]")
                print("    Page d'erreur dans le popup, on recommence...")
                continue
            except Exception:
                pass

            # Vérifier qu'on est bien sur la page SIREN
            wait_and_find(driver, "#saisieSiren", timeout=10)
            print("  Page de saisie SIREN atteinte.")
            return True

        except Exception as e:
            print(f"    Echec tentative {attempt + 1}: {e}")
            time.sleep(ACTION_DELAY * 2)

    raise Exception("Impossible d'atteindre la page de saisie SIREN apres plusieurs tentatives")


def login(driver):
    """
    Attend que l'utilisateur se connecte manuellement sur impots.gouv.fr
    puis navigue vers la page de délégation via les pages intermédiaires.
    """
    driver.get(BASE_URL)
    print("  Connectez-vous manuellement dans le navigateur.")
    input("  Appuyez sur Entree une fois connecte...")
    navigate_to_delegation_page(driver)


# ─── Étapes de délégation ─────────────────────────────────────────────────────

def navigate_to_siren_page(driver):
    """S'assure qu'on est sur la page de saisie SIREN."""
    # Vérifier si on est déjà sur la bonne page
    try:
        driver.find_element(By.CSS_SELECTOR, "#saisieSiren")
    except Exception:
        driver.get(f"{BASE_URL}/opale_usager/SaisieSirenDelegation.do?choixSirenIn=true")
        time.sleep(ACTION_DELAY)
        wait_and_find(driver, "#saisieSiren")


def enter_siren(driver, siren):
    """Entre le SIREN et clique sur Rechercher."""
    fill_input(driver, "#saisieSiren", siren)
    # Le bouton Rechercher est un lien javascript:submitform('saisie')
    driver.execute_script("submitform('saisie');")
    time.sleep(ACTION_DELAY * 2)


def enter_abonne_and_validate(driver, abonne):
    """Entre le numéro d'abonné et clique sur Valider."""
    fill_input(driver, "input[name='num_adh']", abonne)
    wait_and_click(driver, "input[type='submit'][value='Valider']")
    time.sleep(ACTION_DELAY)


def find_service_link(driver, service_label):
    """
    Trouve le lien 'Déléguer ou modifier' correspondant à un service donné.
    On cherche la ligne du tableau qui contient le label du service,
    puis on clique sur le lien 'Déléguer ou modifier' dans la même ligne.
    """
    # Trouver toutes les lignes du tableau des services
    rows = driver.find_elements(By.CSS_SELECTOR, "tr.toutblenc")
    for row in rows:
        labels = row.find_elements(By.CSS_SELECTOR, "label")
        for label in labels:
            if label.text.strip() == service_label:
                # Trouvé ! Cliquer sur le lien "Déléguer ou modifier" de cette ligne
                link = row.find_element(By.CSS_SELECTOR, "a.formLabel")
                return link
    return None


def select_acteur(driver):
    """Sélectionne le radio button 'Acteur' (value='N2') si présent.
    Le name peut être 'role', 'role0', 'role1', etc."""
    try:
        # Chercher tout radio avec value='N2' dont le name commence par 'role'
        radios = driver.find_elements(By.CSS_SELECTOR, "input[type='radio'][value='N2']")
        if radios:
            for radio in radios:
                name = radio.get_attribute("name") or ""
                if name.startswith("role"):
                    if not radio.is_selected():
                        radio.click()
                        time.sleep(ACTION_DELAY / 2)
                    print(f"        Role 'Acteur' selectionne ({name}).")
                    return
        print("        Pas de radio 'Acteur' trouve, on continue.")
    except Exception:
        print("        Pas de choix de role sur cette page, on continue.")


def check_all_checkboxes(driver):
    """Coche toutes les checkboxes sur la page (pour Consulter le Compte fiscal)."""
    checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
    for cb in checkboxes:
        if not cb.is_selected():
            cb.click()
            time.sleep(0.1)
    time.sleep(ACTION_DELAY / 2)


def click_valider(driver):
    """Clique sur le bouton Valider (plusieurs sélecteurs possibles)."""
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "input[type='submit'][value='Valider']")
        btn.click()
    except Exception:
        # Fallback : chercher par XPath tout bouton/input contenant 'Valider'
        btn = driver.find_element(By.XPATH, "//input[@value='Valider']")
        btn.click()
    time.sleep(ACTION_DELAY)


def click_nouvelle_delegation(driver):
    """Clique sur le lien 'Nouvelle délégation'."""
    wait_and_click(driver, "a.lienBlanc[href*='GererDelegation.do']")
    time.sleep(ACTION_DELAY)


def click_nouveau_siren(driver):
    """Clique sur le lien 'Nouveau SIREN'."""
    wait_and_click(driver, "a.lienBlanc[href*='SaisieSirenDelegation.do']")
    time.sleep(ACTION_DELAY)


def process_delegation(driver, abonne, service, is_last=False):
    """
    Effectue une délégation complète pour un service donné.

    Flux :
      1. Sur la page des services, cliquer sur "Déléguer ou modifier" du service
      2. Sélectionner "Acteur" (radio N2)
      3. Si check_all, cocher toutes les checkboxes
      4. Valider
      5. Sur la page récap :
         - Si is_last : cliquer "Nouveau SIREN"
         - Sinon : cliquer "Nouvelle délégation", puis entrer le n° abonné
    """
    service_label = service["label"]
    check_all = service["check_all"]
    print(f"      -> {service_label}...")

    # 1. Cliquer sur "Déléguer ou modifier" pour ce service
    link = find_service_link(driver, service_label)
    if link is None:
        print(f"         Service '{service_label}' non disponible, ignore.")
        return False

    link.click()
    time.sleep(ACTION_DELAY * 2)

    # 2. Sélectionner "Acteur"
    select_acteur(driver)

    # 3. Si consulter compte fiscal, cocher toutes les cases
    if check_all:
        check_all_checkboxes(driver)

    # 4. Valider
    click_valider(driver)

    # 5. Page récap → action suivante
    if is_last:
        click_nouveau_siren(driver)
    else:
        click_nouvelle_delegation(driver)
        enter_abonne_and_validate(driver, abonne)

    return True


def process_siren(driver, siren, abonne):
    """Traite un SIREN complet (délégations disponibles)."""
    print(f"   SIREN {siren}")

    # Naviguer vers la page SIREN et entrer le numéro
    navigate_to_siren_page(driver)
    enter_siren(driver, siren)

    # Entrer le numéro d'abonné et valider
    enter_abonne_and_validate(driver, abonne)

    # Détecter les services disponibles sur la page
    available = []
    for service in SERVICES:
        link = find_service_link(driver, service["label"])
        if link is not None:
            available.append(service)

    if not available:
        print(f"   Aucun service disponible pour SIREN {siren}, passage au suivant.")
        click_nouveau_siren(driver)
        return

    print(f"   {len(available)}/{len(SERVICES)} services disponibles")

    # Effectuer les délégations pour les services disponibles
    for i, service in enumerate(available):
        is_last = (i == len(available) - 1)
        process_delegation(driver, abonne, service, is_last=is_last)

    print(f"   SIREN {siren} termine.")


# ─── Main ─────────────────────────────────────────────────────────────────────

def select_sheets(data):
    """Affiche un menu pour choisir les onglets à traiter."""
    sheets = list(data.keys())
    print("\n  Onglets disponibles :")
    for idx, name in enumerate(sheets, 1):
        nb = len(data[name]["sirens"])
        progress = load_progress(name)
        resume_idx = progress.get("siren_index", 0)
        status = f" (reprise index {resume_idx})" if resume_idx > 0 else ""
        print(f"    {idx}. {name} - {nb} SIRENs{status}")
    print(f"    0. Tous les onglets")

    choix = input("\n  Entrez les numeros des onglets (ex: 1,3,5 ou 0 pour tous) : ").strip()
    if choix == "0":
        return sheets

    selected = []
    for c in choix.split(","):
        c = c.strip()
        if c.isdigit() and 1 <= int(c) <= len(sheets):
            selected.append(sheets[int(c) - 1])
    if not selected:
        print("  Aucun onglet valide selectionne.")
        return []

    print(f"  Onglets selectionnes : {', '.join(selected)}")
    return selected


def main():
    print("=" * 60)
    print("  AUTOMATISATION DELEGATION TVA - impots.gouv.fr")
    print("=" * 60)

    # Lire l'Excel
    data = read_excel()
    if not data:
        print("  Aucune donnee trouvee dans l'Excel.")
        return

    # Sélection des onglets
    selected_sheets = select_sheets(data)
    if not selected_sheets:
        return

    # Initialiser le navigateur
    driver = init_driver()

    try:
        # Connexion manuelle
        login(driver)

        # Boucle sur les onglets sélectionnés
        for sheet_name in selected_sheets:
            sheet_data = data[sheet_name]
            abonne = sheet_data["abonne"]
            sirens = sheet_data["sirens"]

            # Charger la progression pour cet onglet
            progress = load_progress(sheet_name)
            start_index = progress.get("siren_index", 0)

            if start_index > 0:
                print(f"\n  Reprise detectee pour '{sheet_name}' a l'index {start_index}")
                confirm = input("  Reprendre ? (o/n) : ").strip().lower()
                if confirm != "o":
                    clear_progress(sheet_name)
                    start_index = 0

            print(f"\n{'='*60}")
            print(f"  Onglet: {sheet_name} | Abonne: {abonne}")
            print(f"  SIRENs: {len(sirens)} (debut a l'index {start_index})")
            print(f"{'='*60}")

            for i in range(start_index, len(sirens)):
                siren = sirens[i]
                save_progress(sheet_name, i)

                try:
                    process_siren(driver, siren, abonne)
                except Exception as e:
                    print(f"   Erreur sur SIREN {siren} (index {i}): {e}")
                    save_progress(sheet_name, i)
                    retry = input("   Reessayer ? (o/n/q pour quitter) : ").strip().lower()
                    if retry == "q":
                        print("  Arret demande. Progression sauvegardee.")
                        return
                    elif retry == "o":
                        try:
                            process_siren(driver, siren, abonne)
                        except Exception as e2:
                            print(f"   Echec retry: {e2}")
                            print("   Passage au SIREN suivant.")
                            continue
                    else:
                        print("   Passage au SIREN suivant.")
                        continue

                print(f"   Progression: {i + 1}/{len(sirens)} "
                      f"({(i + 1) / len(sirens) * 100:.1f}%)")

            clear_progress(sheet_name)
            print(f"\n  Onglet '{sheet_name}' termine !")

        print("\n" + "=" * 60)
        print("  TOUS LES ONGLETS SELECTIONNES ONT ETE TRAITES !")
        print("=" * 60)

    except KeyboardInterrupt:
        print("\n\n  Interruption manuelle. Progression sauvegardee.")
    except Exception as e:
        print(f"\n  Erreur fatale: {e}")
        print("  Progression sauvegardee.")
    finally:
        input("\nAppuyez sur Entree pour fermer le navigateur...")
        driver.quit()


if __name__ == "__main__":
    main()